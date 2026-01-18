"""
THY Maintenance System - Interactive Excel Simulator
=====================================================
Bu modül, Excel'de interaktif simülatör içeren bir dosya oluşturur.

Özellikler:
- Simülatör sayfası: Dropdown'dan uçak seçimi
- VLOOKUP/INDEX-MATCH formülleri ile otomatik veri çekme
- Koşullu biçimlendirme ile görsel göstergeler
- Progress bar benzeri görselleştirme
"""

import pandas as pd
from datetime import datetime, timedelta
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment


# ============================================
# STYLES
# ============================================

THY_RED = "E31837"
DARK_BLUE = "1a1a2e"
GREEN = "00b894"
YELLOW = "fdcb6e"
ORANGE = "f39c12"
RED = "e74c3c"
LIGHT_BLUE = "74b9ff"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"

header_fill = PatternFill(start_color=THY_RED, end_color=THY_RED, fill_type="solid")
subheader_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
ok_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
warning_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
critical_fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
highlight_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
input_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
output_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")

header_font = Font(bold=True, color="FFFFFF", size=12)
title_font = Font(bold=True, size=18, color=THY_RED)
subtitle_font = Font(bold=True, size=14, color=DARK_BLUE)
label_font = Font(bold=True, size=11)
value_font = Font(size=12)
big_value_font = Font(bold=True, size=16)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

thick_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')


# ============================================
# DATA GENERATION
# ============================================

def generate_fleet_data():
    """THY filo verileri üret"""
    
    fleet_structure = {
        "Airbus A319-100":  {"count": 6,  "prefix": "TC-JL", "category": "NARROW"},
        "Airbus A320-200":  {"count": 14, "prefix": "TC-JP", "category": "NARROW"},
        "Airbus A320 NEO":  {"count": 10, "prefix": "TC-NB", "category": "NARROW"},
        "Airbus A321-200":  {"count": 20, "prefix": "TC-JR", "category": "NARROW"},
        "Airbus A321 NEO":  {"count": 15, "prefix": "TC-LT", "category": "NARROW"},
        "Airbus A330-200":  {"count": 12, "prefix": "TC-JN", "category": "WIDE"},
        "Airbus A330-300":  {"count": 37, "prefix": "TC-JO", "category": "WIDE"},
        "Airbus A350-900":  {"count": 25, "prefix": "TC-LG", "category": "WIDE"},
        "Boeing 737-800":   {"count": 30, "prefix": "TC-JV", "category": "NARROW"},
        "Boeing 737-900ER": {"count": 15, "prefix": "TC-JY", "category": "NARROW"},
        "Boeing 737 MAX 8": {"count": 34, "prefix": "TC-LC", "category": "NARROW"},
        "Boeing 777-300ER": {"count": 34, "prefix": "TC-JJ", "category": "WIDE"},
        "Boeing 787-9":     {"count": 23, "prefix": "TC-LL", "category": "WIDE"},
        "Boeing 777F":      {"count": 8,  "prefix": "TC-LJ", "category": "CARGO"}
    }

    data = []
    today = datetime.now()
    random.seed(42)

    for model, info in fleet_structure.items():
        for i in range(1, info["count"] + 1):
            tail_number = f"{info['prefix']}{chr(65 + (i % 26))}{random.randint(10, 99)}"
            
            if info["category"] in ["WIDE", "CARGO"]:
                total_fh = random.randint(5000, 50000)
                avg_flight_time = random.uniform(5, 10)
            else:
                total_fh = random.randint(2000, 35000)
                avg_flight_time = random.uniform(1.5, 3.5)
            
            total_fc = int(total_fh / avg_flight_time)
            years_in_service = random.randint(2, 15)
            delivery_date = today - timedelta(days=years_in_service * 365)
            
            check_types = ["A", "B", "C"]
            last_check = random.choice(check_types)
            
            if last_check == "A":
                fh_since_check = random.randint(50, 580)
                fc_since_check = random.randint(30, 380)
            elif last_check == "B":
                fh_since_check = random.randint(100, 2000)
                fc_since_check = random.randint(80, 1200)
            else:
                fh_since_check = random.randint(500, 5500)
                fc_since_check = random.randint(300, 3500)
            
            last_maint_date = today - timedelta(days=random.randint(10, 600))
            years_since_d_check = random.randint(0, 6)
            last_d_check = today - timedelta(days=years_since_d_check * 365 + random.randint(0, 180))
            daily_fh = round(random.uniform(6, 14), 1)
            status = random.choice(["Aktif", "Aktif", "Aktif", "Bakımda"])
            
            days_since_maint = (today - last_maint_date).days
            days_since_d = (today - last_d_check).days
            
            data.append({
                "Kuyruk No": tail_number,
                "Model": model,
                "Kategori": info["category"],
                "Teslim Tarihi": delivery_date.strftime("%Y-%m-%d"),
                "Toplam FH": total_fh,
                "Toplam FC": total_fc,
                "Son Bakım Tipi": last_check,
                "Son Bakımdan Beri FH": fh_since_check,
                "Son Bakımdan Beri FC": fc_since_check,
                "Son Bakım Tarihi": last_maint_date.strftime("%Y-%m-%d"),
                "Son Bakımdan Beri Gün": days_since_maint,
                "Son D-Check": last_d_check.strftime("%Y-%m-%d"),
                "Son D-Check Beri Gün": days_since_d,
                "Günlük Ort. FH": daily_fh,
                "Durum": status
            })

    return pd.DataFrame(data).sort_values("Kuyruk No").reset_index(drop=True)


# ============================================
# EXCEL CREATION
# ============================================

def create_simulator_excel(df):
    """Interaktif simülatör Excel dosyası oluştur"""
    
    wb = Workbook()
    
    # ========== SHEET 1: SIMULATOR ==========
    ws_sim = wb.active
    ws_sim.title = "Simülatör"
    create_simulator_sheet(ws_sim, df)
    
    # ========== SHEET 2: DATABASE ==========
    ws_db = wb.create_sheet("Veritabanı")
    create_database_sheet(ws_db, df)
    
    # ========== SHEET 3: MAINTENANCE RULES ==========
    ws_rules = wb.create_sheet("Bakım Kuralları")
    create_rules_sheet(ws_rules)
    
    # ========== SHEET 4: REFERENCES ==========
    ws_refs = wb.create_sheet("Akademik Referanslar")
    create_references_sheet(ws_refs)
    
    return wb


def create_simulator_sheet(ws, df):
    """Interaktif simülatör sayfası"""
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 5
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 20
    
    # ========== HEADER ==========
    ws.merge_cells('B2:K2')
    ws['B2'] = "✈️ THY AIRCRAFT MAINTENANCE SIMULATOR"
    ws['B2'].font = Font(bold=True, size=22, color=THY_RED)
    ws['B2'].alignment = center_align
    
    ws.merge_cells('B3:K3')
    ws['B3'] = "Uçak Bakım Karar Destek Sistemi - Excel Simülasyonu"
    ws['B3'].font = Font(size=12, italic=True)
    ws['B3'].alignment = center_align
    
    ws.merge_cells('B4:K4')
    ws['B4'] = f"Sistem Tarihi: {datetime.now().strftime('%Y-%m-%d')} | Referanslar: Papakostas (2010), Callewaert (2017), Kowalski (2021)"
    ws['B4'].font = Font(size=10, color="666666")
    ws['B4'].alignment = center_align
    
    # ========== INPUT SECTION ==========
    ws.merge_cells('B6:C6')
    ws['B6'] = "📋 UÇAK SEÇİMİ"
    ws['B6'].font = header_font
    ws['B6'].fill = header_fill
    ws['B6'].alignment = center_align
    
    # Model Selection
    ws['B8'] = "Uçak Modeli:"
    ws['B8'].font = label_font
    ws['C8'] = "Boeing 777-300ER"  # Default value
    ws['C8'].fill = input_fill
    ws['C8'].border = thick_border
    ws['C8'].font = value_font
    
    # Create dropdown for models
    models = sorted(df["Model"].unique().tolist())
    model_validation = DataValidation(
        type="list",
        formula1=f'"' + ','.join(models) + '"',
        allow_blank=False
    )
    model_validation.error = "Lütfen listeden bir model seçin"
    model_validation.errorTitle = "Geçersiz Seçim"
    model_validation.prompt = "Uçak modelini seçin"
    model_validation.promptTitle = "Model Seçimi"
    ws.add_data_validation(model_validation)
    model_validation.add(ws['C8'])
    
    # Tail Number Selection
    ws['B10'] = "Kuyruk Numarası:"
    ws['B10'].font = label_font
    ws['C10'] = "TC-JJA10"  # Default - will be updated by formula
    ws['C10'].fill = input_fill
    ws['C10'].border = thick_border
    ws['C10'].font = value_font
    
    # Note: In real Excel, this would be a dynamic dropdown based on model
    # For now, include all tail numbers
    tails = sorted(df["Kuyruk No"].unique().tolist())
    tail_validation = DataValidation(
        type="list",
        formula1=f'"' + ','.join(tails[:100]) + '"',  # Limit to first 100 for Excel limits
        allow_blank=False
    )
    tail_validation.error = "Lütfen listeden bir kuyruk numarası seçin"
    ws.add_data_validation(tail_validation)
    tail_validation.add(ws['C10'])
    
    # Instructions
    ws['B12'] = "💡 Kullanım:"
    ws['B12'].font = Font(bold=True, color=DARK_BLUE)
    ws['B13'] = "1. Yukarıdaki dropdown'lardan uçak seçin"
    ws['B14'] = "2. Sağ taraftaki bakım durumu otomatik güncellenecek"
    ws['B15'] = "3. Veritabanı sayfasından tüm uçakları görüntüleyebilirsiniz"
    
    # ========== AIRCRAFT INFO SECTION ==========
    ws.merge_cells('E6:G6')
    ws['E6'] = "🛫 UÇAK BİLGİLERİ"
    ws['E6'].font = header_font
    ws['E6'].fill = subheader_fill
    ws['E6'].alignment = center_align
    
    # Info rows with VLOOKUP formulas (simulated with static data for now)
    info_labels = [
        ("Model", 8),
        ("Kategori", 9),
        ("Toplam Uçuş Saati (FH)", 10),
        ("Toplam Döngü (FC)", 11),
        ("Son Bakım Tipi", 12),
        ("Son Bakımdan Beri FH", 13),
        ("Son Bakımdan Beri FC", 14),
        ("Son Bakım Tarihi", 15),
        ("Günlük Ort. FH", 16),
        ("Mevcut Durum", 17)
    ]
    
    for label, row in info_labels:
        ws[f'E{row}'] = label + ":"
        ws[f'E{row}'].font = label_font
        ws[f'E{row}'].border = thin_border
        
        # Formula to lookup from database
        col_index = ["Kuyruk No", "Model", "Kategori", "Teslim Tarihi", "Toplam FH", 
                     "Toplam FC", "Son Bakım Tipi", "Son Bakımdan Beri FH", 
                     "Son Bakımdan Beri FC", "Son Bakım Tarihi", "Son Bakımdan Beri Gün",
                     "Son D-Check", "Son D-Check Beri Gün", "Günlük Ort. FH", "Durum"]
        
        # Map label to column index
        label_to_col = {
            "Model": 2,
            "Kategori": 3,
            "Toplam Uçuş Saati (FH)": 5,
            "Toplam Döngü (FC)": 6,
            "Son Bakım Tipi": 7,
            "Son Bakımdan Beri FH": 8,
            "Son Bakımdan Beri FC": 9,
            "Son Bakım Tarihi": 10,
            "Günlük Ort. FH": 14,
            "Mevcut Durum": 15
        }
        
        col_idx = label_to_col.get(label, 2)
        ws[f'F{row}'] = f'=IFERROR(VLOOKUP($C$10,Veritabanı!$A$3:$O$290,{col_idx},FALSE),"-")'
        ws[f'F{row}'].fill = output_fill
        ws[f'F{row}'].border = thin_border
        ws[f'F{row}'].font = value_font
    
    # ========== MAINTENANCE STATUS SECTION ==========
    ws.merge_cells('J6:K6')
    ws['J6'] = "⚙️ BAKIM DURUMU"
    ws['J6'].font = header_font
    ws['J6'].fill = header_fill
    ws['J6'].alignment = center_align
    
    # Maintenance calculations
    maint_checks = [
        ("A Check", 8, "FH", 600, 400),
        ("B Check (Phased)", 12, "Days", 180, None),
        ("C Check", 16, "FH", 6000, 730),
        ("D Check (Heavy)", 20, "Days", 2190, None)
    ]
    
    for check_name, start_row, check_type, limit1, limit2 in maint_checks:
        # Check name header
        ws.merge_cells(f'J{start_row}:K{start_row}')
        ws[f'J{start_row}'] = check_name
        ws[f'J{start_row}'].font = Font(bold=True, size=12)
        ws[f'J{start_row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws[f'J{start_row}'].border = thick_border
        ws[f'J{start_row}'].alignment = center_align
        
        # Progress percentage
        ws[f'J{start_row+1}'] = "İlerleme (%):"
        ws[f'J{start_row+1}'].font = label_font
        
        if check_name == "A Check":
            # A Check: MAX of FH/600 and FC/400
            ws[f'K{start_row+1}'] = f'=IFERROR(MIN(100,MAX(VLOOKUP($C$10,Veritabanı!$A$3:$O$290,8,FALSE)/{limit1}*100,VLOOKUP($C$10,Veritabanı!$A$3:$O$290,9,FALSE)/{limit2}*100)),0)'
        elif check_name == "B Check (Phased)":
            # B Check: Days since last maintenance / 180
            ws[f'K{start_row+1}'] = f'=IFERROR(MIN(100,VLOOKUP($C$10,Veritabanı!$A$3:$O$290,11,FALSE)/{limit1}*100),0)'
        elif check_name == "C Check":
            # C Check: MAX of FH/6000 and Days/730
            ws[f'K{start_row+1}'] = f'=IFERROR(MIN(100,MAX(VLOOKUP($C$10,Veritabanı!$A$3:$O$290,8,FALSE)*2/{limit1}*100,VLOOKUP($C$10,Veritabanı!$A$3:$O$290,11,FALSE)/{limit2}*100)),0)'
        else:  # D Check
            # D Check: Days since D-Check / 2190
            ws[f'K{start_row+1}'] = f'=IFERROR(MIN(100,VLOOKUP($C$10,Veritabanı!$A$3:$O$290,13,FALSE)/{limit1}*100),0)'
        
        ws[f'K{start_row+1}'].fill = output_fill
        ws[f'K{start_row+1}'].border = thin_border
        ws[f'K{start_row+1}'].number_format = '0.0"%"'
        
        # Status
        ws[f'J{start_row+2}'] = "Durum:"
        ws[f'J{start_row+2}'].font = label_font
        ws[f'K{start_row+2}'] = f'=IF(K{start_row+1}>=90,"🔴 KRİTİK",IF(K{start_row+1}>=75,"🟡 UYARI","🟢 NORMAL"))'
        ws[f'K{start_row+2}'].border = thin_border
        ws[f'K{start_row+2}'].font = Font(bold=True, size=11)
        
        # Add conditional formatting for status
        ws.conditional_formatting.add(
            f'K{start_row+2}',
            FormulaRule(
                formula=[f'K{start_row+1}>=90'],
                fill=critical_fill
            )
        )
        ws.conditional_formatting.add(
            f'K{start_row+2}',
            FormulaRule(
                formula=[f'AND(K{start_row+1}>=75,K{start_row+1}<90)'],
                fill=warning_fill
            )
        )
        ws.conditional_formatting.add(
            f'K{start_row+2}',
            FormulaRule(
                formula=[f'K{start_row+1}<75'],
                fill=ok_fill
            )
        )
    
    # ========== SUMMARY BOX ==========
    ws.merge_cells('B20:C20')
    ws['B20'] = "📊 EN KRİTİK BAKIM"
    ws['B20'].font = header_font
    ws['B20'].fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    ws['B20'].alignment = center_align
    
    ws['B21'] = "Bakım Tipi:"
    ws['B21'].font = label_font
    ws['C21'] = '=IF(MAX(K9,K13,K17,K21)=K9,"A Check",IF(MAX(K9,K13,K17,K21)=K13,"B Check",IF(MAX(K9,K13,K17,K21)=K17,"C Check","D Check")))'
    ws['C21'].font = big_value_font
    ws['C21'].fill = output_fill
    ws['C21'].border = thick_border
    
    ws['B22'] = "İlerleme:"
    ws['B22'].font = label_font
    ws['C22'] = '=MAX(K9,K13,K17,K21)'
    ws['C22'].font = big_value_font
    ws['C22'].fill = output_fill
    ws['C22'].border = thick_border
    ws['C22'].number_format = '0.0"%"'
    
    ws['B23'] = "Genel Durum:"
    ws['B23'].font = label_font
    ws['C23'] = '=IF(C22>=90,"🔴 ACİL BAKIM GEREKLİ!",IF(C22>=75,"🟡 BAKIM YAKLAŞIYOR","🟢 NORMAL"))'
    ws['C23'].font = big_value_font
    ws['C23'].border = thick_border
    
    # Conditional formatting for overall status
    ws.conditional_formatting.add(
        'C23',
        FormulaRule(formula=['C22>=90'], fill=critical_fill)
    )
    ws.conditional_formatting.add(
        'C23',
        FormulaRule(formula=['AND(C22>=75,C22<90)'], fill=warning_fill)
    )
    ws.conditional_formatting.add(
        'C23',
        FormulaRule(formula=['C22<75'], fill=ok_fill)
    )
    
    # ========== LEGEND ==========
    ws['B26'] = "📖 RENK KODLARI:"
    ws['B26'].font = Font(bold=True)
    
    ws['B27'] = "🔴 KRİTİK"
    ws['B27'].fill = critical_fill
    ws['C27'] = "≥90% - Acil bakım planlanmalı"
    
    ws['B28'] = "🟡 UYARI"
    ws['B28'].fill = warning_fill
    ws['C28'] = "75-89% - Bakım penceresi yaklaşıyor"
    
    ws['B29'] = "🟢 NORMAL"
    ws['B29'].fill = ok_fill
    ws['C29'] = "<75% - Normal operasyon devam"
    
    # Row heights
    ws.row_dimensions[2].height = 35
    ws.row_dimensions[6].height = 25


def create_database_sheet(ws, df):
    """Veritabanı sayfası (lookup için)"""
    
    ws.merge_cells('A1:O1')
    ws['A1'] = "📊 THY FİLO VERİTABANI - VLOOKUP KAYNAK TABLOSU"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    # Headers
    columns = ["Kuyruk No", "Model", "Kategori", "Teslim Tarihi", "Toplam FH", 
               "Toplam FC", "Son Bakım Tipi", "Son Bakımdan Beri FH", 
               "Son Bakımdan Beri FC", "Son Bakım Tarihi", "Son Bakımdan Beri Gün",
               "Son D-Check", "Son D-Check Beri Gün", "Günlük Ort. FH", "Durum"]
    
    for i, col in enumerate(columns):
        cell = ws.cell(row=2, column=i+1)
        cell.value = col
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # Data
    for r_idx, row in df.iterrows():
        for c_idx, col in enumerate(columns):
            cell = ws.cell(row=r_idx+3, column=c_idx+1)
            cell.value = row[col]
            cell.border = thin_border
            cell.alignment = center_align
    
    # Column widths
    col_widths = [12, 20, 10, 12, 12, 12, 12, 18, 18, 15, 18, 12, 18, 12, 10]
    for i, width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = width
    
    # Freeze panes
    ws.freeze_panes = 'A3'


def create_rules_sheet(ws):
    """Bakım kuralları sayfası"""
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "📋 BAKIM LİMİTLERİ VE KURALLARI (EASA/FAA Standartları)"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    # Headers
    headers = ["Bakım Tipi", "FH Limiti", "FC Limiti", "Zaman Limiti", "Tahmini Süre"]
    for i, header in enumerate(headers):
        cell = ws.cell(row=3, column=i+1)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # Data
    rules = [
        ("A Check", 600, 400, "-", "1 gün (24 saat)"),
        ("B Check (Phased Maintenance)", "-", "-", "180 gün (6 ay)", "3 gün (72 saat)"),
        ("C Check", 6000, "-", "730 gün (2 yıl)", "7 gün (168 saat)"),
        ("D Check (Heavy Maintenance)", "-", "-", "2190 gün (6 yıl)", "30 gün (720 saat)")
    ]
    
    for r_idx, row_data in enumerate(rules):
        for c_idx, value in enumerate(row_data):
            cell = ws.cell(row=r_idx+4, column=c_idx+1)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_align
    
    # Formulas explanation
    ws['A9'] = "📐 HESAPLAMA FORMÜLLERİ:"
    ws['A9'].font = Font(bold=True, size=14)
    
    formulas = [
        ("A Check İlerleme (%)", "MAX(Son_Bakımdan_Beri_FH / 600 , Son_Bakımdan_Beri_FC / 400) * 100"),
        ("B Check İlerleme (%)", "Son_Bakımdan_Beri_Gün / 180 * 100"),
        ("C Check İlerleme (%)", "MAX(Son_Bakımdan_Beri_FH * 2 / 6000 , Son_Bakımdan_Beri_Gün / 730) * 100"),
        ("D Check İlerleme (%)", "Son_D_Check_Beri_Gün / 2190 * 100"),
        ("Durum", 'IF(İlerleme >= 90, "KRİTİK", IF(İlerleme >= 75, "UYARI", "NORMAL"))')
    ]
    
    for i, (name, formula) in enumerate(formulas):
        ws[f'A{11+i}'] = name
        ws[f'A{11+i}'].font = label_font
        ws[f'B{11+i}'] = formula
        ws[f'B{11+i}'].font = Font(name='Consolas', size=10)
        ws.merge_cells(f'B{11+i}:E{11+i}')
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20


def create_references_sheet(ws):
    """Akademik referanslar sayfası"""
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "📚 AKADEMİK REFERANSLAR"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    references = [
        ("Papakostas et al. (2010)", 
         "Operational Aircraft Maintenance Planning: A Multi-objective Approach",
         "B Check modern havacılıkta 'Phased Maintenance' olarak uygulanır."),
        
        ("Callewaert et al. (2017)", 
         "Integrating maintenance work progress monitoring in a stochastic framework",
         "Bakımların %15'inde Non-Routine Finding çıkar, süre 1-3 gün uzar."),
        
        ("Kowalski et al. (2021)", 
         "Resource-constrained project scheduling for aircraft maintenance",
         "Hangar kapasitesi sınırlıdır, eş zamanlı max 5 geniş gövde bakımı."),
        
        ("Hollander (2025)", 
         "Uncertainty Quantification in Aviation Maintenance Planning",
         "Belirsizlik olasılık dağılımlarıyla modellenmelidir.")
    ]
    
    row = 3
    for author, title, note in references:
        ws[f'A{row}'] = author
        ws[f'A{row}'].font = Font(bold=True, size=12, color=THY_RED)
        
        ws[f'A{row+1}'] = f"📄 {title}"
        ws[f'A{row+1}'].font = Font(italic=True)
        
        ws[f'A{row+2}'] = f"💡 {note}"
        ws[f'A{row+2}'].font = Font(color="0066CC")
        
        row += 4
    
    ws.column_dimensions['A'].width = 100


def main():
    """Ana fonksiyon"""
    
    print("=" * 60)
    print("THY Maintenance System - Interactive Excel Simulator")
    print("=" * 60)
    
    # Generate data
    print("\n📊 Filo verileri oluşturuluyor...")
    df = generate_fleet_data()
    print(f"   ✓ {len(df)} uçak verisi üretildi")
    
    # Create workbook
    print("\n📑 Excel simülatörü oluşturuluyor...")
    wb = create_simulator_excel(df)
    
    # Save
    output_path = "THY_Maintenance_Simulator.xlsx"
    wb.save(output_path)
    print(f"   ✓ Dosya kaydedildi: {output_path}")
    
    print("\n" + "=" * 60)
    print("✅ Excel Simülatörü başarıyla oluşturuldu!")
    print("=" * 60)
    
    print("\n📋 Sayfalar:")
    print("   1. Simülatör - Dropdown'dan uçak seç, bakım durumunu gör")
    print("   2. Veritabanı - Tüm filo verileri (VLOOKUP kaynağı)")
    print("   3. Bakım Kuralları - A/B/C/D check limitleri")
    print("   4. Akademik Referanslar - Literatür")
    
    print("\n🎯 Kullanım:")
    print("   1. Simülatör sayfasını aç")
    print("   2. C8 hücresinden uçak modeli seç")
    print("   3. C10 hücresinden kuyruk numarası seç")
    print("   4. Sağ taraftaki bakım durumu otomatik güncellenir!")
    
    return output_path


if __name__ == "__main__":
    main()
